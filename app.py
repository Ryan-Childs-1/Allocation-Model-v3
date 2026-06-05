
import io
import json
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st

# Optional Excel writer/reader helpers
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except Exception:
    load_workbook = None


# ============================================================
# Base Model v1/v2 Allocation Streamlit App
# ------------------------------------------------------------
# Purpose:
# - Upload allocation workbook/CSV.
# - Select Base Model v1 or Base Model v2.
# - Fill Final Alloc. only.
# - Preserve blanks where rows should not be allocated.
# - Return a downloadable result file plus an audit summary.
#
# Key business assumptions from allocation AI design:
# - BR = Flag
# - BS = Final Alloc.
# - BN = Alloc. Rec.
# - BM = Proj. Demand
# - BT = Left DC / remaining units
# - BU = Final Supply
# - BK = FLM
# - AP = QOH
# - AQ = Supply
# - AK/AL/AM/AN/AO = L30/D30/D60/LW/TTM
# - Include Allocate, Review, and Z - No Alloc rows as candidates.
# - Use three passes.
# - No hard cap on number of FLMs per pass or total.
# - Round to FLM, but allow remainder allocation when Left DC is below one FLM.
# - Blank FLM defaults to 1.
# ============================================================


st.set_page_config(
    page_title="Allocation AI — Base Model v1/v2",
    page_icon="📦",
    layout="wide",
)

APP_VERSION = "Base Model v1/v2 Streamlit App — Completed"
DEFAULT_SHEET_INDEX = 0


COLUMN_LETTERS = {
    "O": 15,
    "AK": 37,
    "AL": 38,
    "AM": 39,
    "AN": 40,
    "AO": 41,
    "AP": 42,
    "AQ": 43,
    "BJ": 62,
    "BK": 63,
    "BM": 65,
    "BN": 66,
    "BR": 70,
    "BS": 71,
    "BT": 72,
    "BU": 73,
    "CA": 79,
    "CB": 80,
}


SYNONYMS = {
    "group_key": ["Matching Group", "Matching group", "Group", "Item", "UPC", "Line Id", "Line ID", "SKU", "O"],
    "l30": ["L30", "L30 Sales", "Last 30", "Last 30 Days", "AK"],
    "d30": ["D30", "D30 Demand", "Demand 30", "AL"],
    "d60": ["D60", "Demand 60", "60 Day Demand", "AM"],
    "lw": ["LW", "LW Sales", "Last Week", "AN"],
    "ttm": ["TTM", "TTM Sales", "Trailing 12", "AO"],
    "qoh": ["QOH", "Qty On Hand", "Quantity on Hand", "AP"],
    "supply": ["Supply", "Supply on Hand", "SOH", "AQ"],
    "mil": ["MIL", "Min Inv Level", "Minimum Inventory Level", "BJ"],
    "flm": ["FLM", "Allocation Unit", "Pack", "Pack Multiple", "BK"],
    "proj_demand": ["Proj. Demand", "Projected Demand", "Proj Demand", "BM"],
    "alloc_rec": ["Alloc. Rec.", "Alloc Rec", "Allocation Recommendation", "BN"],
    "flag": ["Flag", "Review Flag", "BR"],
    "final_alloc": ["Final Alloc.", "Final Alloc", "Final Allocation", "BS"],
    "left_dc": ["Left DC", "Left in DC", "Left-in-DC", "DC Avail", "Remaining DC", "BT"],
    "final_supply": ["Final Supply", "Final Supply Units", "BU"],
    "demand_discount": ["Demand Discount", "Demand Disc", "CA"],
    "base_demand": ["New Base Demand", "Base Demand", "CB"],
}


@dataclass
class ModelConfig:
    name: str
    description: str
    demand_weight: float
    alloc_rec_weight: float
    min_need_weight: float
    review_multiplier: float
    z_no_alloc_multiplier: float
    safety_stock_weight: float
    overstock_buffer_flm: float
    demand_discount_weight: float
    pass_multipliers: Tuple[float, float, float]


MODEL_CONFIGS = {
    "Base Model v1": ModelConfig(
        name="Base Model v1",
        description=(
            "Conservative baseline built from the Version 6 logic. "
            "Uses projected demand, allocation recommendation, current supply, FLM rounding, "
            "and demand-protective caps."
        ),
        demand_weight=0.74,
        alloc_rec_weight=0.68,
        min_need_weight=0.30,
        review_multiplier=0.55,
        z_no_alloc_multiplier=0.35,
        safety_stock_weight=0.18,
        overstock_buffer_flm=1.00,
        demand_discount_weight=0.35,
        pass_multipliers=(0.70, 0.90, 1.00),
    ),
    "Base Model v2": ModelConfig(
        name="Base Model v2",
        description=(
            "Improved baseline built from the Version 8 design. "
            "More advanced demand blend, stronger use of BM/BN, includes Z - No Alloc opportunity rows, "
            "and is more willing to allocate when demand support is broad."
        ),
        demand_weight=0.92,
        alloc_rec_weight=0.82,
        min_need_weight=0.42,
        review_multiplier=0.72,
        z_no_alloc_multiplier=0.52,
        safety_stock_weight=0.26,
        overstock_buffer_flm=1.15,
        demand_discount_weight=0.50,
        pass_multipliers=(0.85, 1.05, 1.20),
    ),
}


def excel_col_name(n: int) -> str:
    name = ""
    while n:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name


def normalize_name(x) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(x).strip().lower())


def to_number(series_or_value, default=0.0):
    if isinstance(series_or_value, pd.Series):
        s = series_or_value.copy()
        s = s.replace({"-": np.nan, "": np.nan, " ": np.nan})
        if s.dtype == object:
            s = s.astype(str).str.replace("$", "", regex=False).str.replace(",", "", regex=False).str.replace("%", "", regex=False)
        return pd.to_numeric(s, errors="coerce").fillna(default)
    try:
        if pd.isna(series_or_value):
            return default
        if isinstance(series_or_value, str):
            clean = series_or_value.replace("$", "").replace(",", "").replace("%", "").strip()
            if clean in ("", "-"):
                return default
            return float(clean)
        return float(series_or_value)
    except Exception:
        return default


def safe_str_series(s: pd.Series) -> pd.Series:
    return s.fillna("").astype(str)


def find_col(df: pd.DataFrame, key: str) -> Optional[str]:
    cols = list(df.columns)
    normalized_cols = {normalize_name(c): c for c in cols}

    # 1) Synonym match
    for cand in SYNONYMS.get(key, []):
        n = normalize_name(cand)
        if n in normalized_cols:
            return normalized_cols[n]

    # 2) Partial match
    for cand in SYNONYMS.get(key, []):
        nc = normalize_name(cand)
        for c in cols:
            if nc and (nc in normalize_name(c) or normalize_name(c) in nc):
                return c

    # 3) Excel letter fallback by physical position
    for cand in SYNONYMS.get(key, []):
        cand_upper = str(cand).upper()
        if cand_upper in COLUMN_LETTERS:
            pos = COLUMN_LETTERS[cand_upper] - 1
            if 0 <= pos < len(cols):
                return cols[pos]

    return None


def build_column_map(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    return {key: find_col(df, key) for key in SYNONYMS.keys()}


def read_uploaded_file(uploaded_file, sheet_name=None) -> Tuple[pd.DataFrame, Dict]:
    name = uploaded_file.name.lower()
    raw = uploaded_file.read()
    meta = {"filename": uploaded_file.name, "input_type": None, "sheet_name": sheet_name}

    if name.endswith(".csv"):
        meta["input_type"] = "csv"
        df = pd.read_csv(io.BytesIO(raw), dtype=object)
        return df, meta

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        meta["input_type"] = "xlsx"
        xl = pd.ExcelFile(io.BytesIO(raw), engine="openpyxl")
        use_sheet = sheet_name if sheet_name is not None else xl.sheet_names[DEFAULT_SHEET_INDEX]
        meta["sheet_name"] = use_sheet
        df = pd.read_excel(io.BytesIO(raw), sheet_name=use_sheet, dtype=object, engine="openpyxl")
        meta["raw_bytes"] = raw
        return df, meta

    if name.endswith(".xlsb"):
        meta["input_type"] = "xlsb"
        # pyxlsb is required by requirements.txt. Streamlit Cloud will install it.
        xl = pd.ExcelFile(io.BytesIO(raw), engine="pyxlsb")
        use_sheet = sheet_name if sheet_name is not None else xl.sheet_names[DEFAULT_SHEET_INDEX]
        meta["sheet_name"] = use_sheet
        df = pd.read_excel(io.BytesIO(raw), sheet_name=use_sheet, dtype=object, engine="pyxlsb")
        return df, meta

    raise ValueError("Unsupported file type. Upload CSV, XLSX, XLSM, or XLSB.")


def get_sheet_names(uploaded_file) -> List[str]:
    name = uploaded_file.name.lower()
    raw = uploaded_file.getvalue()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return pd.ExcelFile(io.BytesIO(raw), engine="openpyxl").sheet_names
    if name.endswith(".xlsb"):
        return pd.ExcelFile(io.BytesIO(raw), engine="pyxlsb").sheet_names
    return []


def mround_to_flm(value: float, flm: float, left: float) -> int:
    value = max(0.0, float(value))
    left = max(0.0, float(left))
    flm = max(1.0, float(flm))

    if left <= 0 or value <= 0:
        return 0

    # Critical business rule: if remaining DC is below one FLM, allow remaining units.
    if left < flm:
        return int(math.floor(left))

    rounded = int(round(value / flm) * flm)
    if rounded <= 0 and value >= max(1, flm * 0.35):
        rounded = int(flm)

    rounded = min(rounded, int(math.floor(left)))
    if rounded >= flm:
        rounded = int(math.floor(rounded / flm) * flm)
    return max(0, int(rounded))


def demand_blend(row, cmap, cfg: ModelConfig) -> float:
    def val(k):
        col = cmap.get(k)
        return to_number(row[col]) if col and col in row.index else 0.0

    l30 = val("l30")
    d30 = val("d30")
    d60 = val("d60")
    lw = val("lw")
    ttm = val("ttm")
    bm = val("proj_demand")
    bn = val("alloc_rec")
    mil = val("mil")
    ca = val("demand_discount")
    cb = val("base_demand")

    components = [
        l30 * 1.18,
        d30 * 0.90,
        d60 * 0.55,
        lw * 4.29,
        ttm / 12.0,
        bm * 1.00 if bm > 0 else 0,
        cb * 1.00 if cb > 0 else 0,
    ]
    positive = [x for x in components if x and x > 0]
    if positive:
        # Weighted robust demand: use both median and high percentile to avoid one noisy metric dominating.
        median_part = float(np.median(positive))
        high_part = float(np.percentile(positive, 70))
        raw_demand = 0.55 * median_part + 0.45 * high_part
    else:
        raw_demand = 0.0

    if bm > 0:
        raw_demand = max(raw_demand, bm * cfg.demand_weight)

    # Demand discount is typically a decimal that downweights underperforming stores.
    # Treat 0/blank as neutral because historical files often leave it blank.
    if ca > 0:
        if ca <= 1.5:
            raw_demand *= max(0.45, 1.0 - cfg.demand_discount_weight * max(0.0, 1.0 - ca))
        else:
            raw_demand *= max(0.45, min(1.15, ca / 100.0))

    # BN is a strong signal but not allowed to blindly override demand protection.
    if bn > 0:
        raw_demand = max(raw_demand, bn * cfg.alloc_rec_weight)

    # MIL gets partial weight only when there is some real demand signal.
    if mil > 0 and (l30 + d30 + d60 + lw + ttm + bm + cb) > 0:
        raw_demand = max(raw_demand, mil * cfg.min_need_weight)

    return max(0.0, raw_demand)


def row_candidate_status(flag_text: str) -> str:
    f = str(flag_text or "").upper()
    if "ALLOC" in f:
        return "ALLOCATE"
    if "REVIEW" in f:
        return "REVIEW"
    if "Z" in f and "NO" in f and "ALLOC" in f:
        return "Z_NO_ALLOC"
    if "NO ALLOC" in f:
        return "Z_NO_ALLOC"
    return "IGNORE"


def compute_model_predictions(df: pd.DataFrame, cmap: Dict[str, Optional[str]], cfg: ModelConfig) -> Tuple[pd.DataFrame, pd.DataFrame]:
    out = df.copy()
    n = len(out)
    pred = np.zeros(n, dtype=int)
    pass_allocs = {1: np.zeros(n, dtype=int), 2: np.zeros(n, dtype=int), 3: np.zeros(n, dtype=int)}
    reasons = [""] * n

    # Required-ish columns with fallbacks
    flag_col = cmap.get("flag")
    final_col = cmap.get("final_alloc")
    left_col = cmap.get("left_dc")
    flm_col = cmap.get("flm")
    supply_col = cmap.get("supply")
    qoh_col = cmap.get("qoh")
    final_supply_col = cmap.get("final_supply")
    d60_col = cmap.get("d60")
    ttm_col = cmap.get("ttm")
    d30_col = cmap.get("d30")
    l30_col = cmap.get("l30")
    group_col = cmap.get("group_key")

    flags = safe_str_series(out[flag_col]) if flag_col else pd.Series([""] * n)
    left_dc = to_number(out[left_col], 0) if left_col else pd.Series([0] * n)
    flm = to_number(out[flm_col], 1) if flm_col else pd.Series([1] * n)
    flm = flm.where(flm > 0, 1)

    qoh = to_number(out[qoh_col], 0) if qoh_col else pd.Series([0] * n)
    supply = to_number(out[supply_col], 0) if supply_col else pd.Series([0] * n)
    d60 = to_number(out[d60_col], 0) if d60_col else pd.Series([0] * n)
    ttm = to_number(out[ttm_col], 0) if ttm_col else pd.Series([0] * n)
    d30 = to_number(out[d30_col], 0) if d30_col else pd.Series([0] * n)
    l30 = to_number(out[l30_col], 0) if l30_col else pd.Series([0] * n)

    if group_col:
        groups = safe_str_series(out[group_col]).replace("", "__NO_GROUP__")
    else:
        groups = pd.Series([f"row_{i}" for i in range(n)])

    # Use max positive Left DC by group as available pool. This handles files where BT repeats per item/store row.
    group_left = {}
    for g, vals in left_dc.groupby(groups):
        max_left = float(np.nanmax(vals.values)) if len(vals) else 0.0
        group_left[g] = max(0.0, max_left)

    # Rank rows inside each group by need so scarce DC units go to strongest stores first.
    scoring_rows = []
    raw_need = np.zeros(n, dtype=float)
    status_list = []
    for i, row in out.iterrows():
        status = row_candidate_status(flags.iloc[i])
        status_list.append(status)

        if status == "IGNORE":
            raw_need[i] = 0.0
            continue

        base_demand = demand_blend(row, cmap, cfg)
        current_supply = max(0.0, qoh.iloc[i] + supply.iloc[i])
        effective_need = max(0.0, base_demand - current_supply)

        if status == "REVIEW":
            effective_need *= cfg.review_multiplier
        elif status == "Z_NO_ALLOC":
            effective_need *= cfg.z_no_alloc_multiplier

        # Mild safety stock if there is real recent demand and the store is at/near zero.
        if current_supply <= 0 and (l30.iloc[i] + d30.iloc[i] + d60.iloc[i]) > 0:
            effective_need += max(1.0, flm.iloc[i] * cfg.safety_stock_weight)

        raw_need[i] = effective_need
        scoring_rows.append((groups.iloc[i], i, effective_need))

    # Sort per group by need descending, but keep stable row order as tie-breaker.
    sorted_indices_by_group = {}
    for g, i, need in scoring_rows:
        sorted_indices_by_group.setdefault(g, []).append((i, need))
    for g in sorted_indices_by_group:
        sorted_indices_by_group[g].sort(key=lambda x: (-x[1], x[0]))

    # Three model passes. Each pass may add; no FLM cap besides DC availability and demand protection.
    for pass_num, pass_multiplier in enumerate(cfg.pass_multipliers, start=1):
        for g, pairs in sorted_indices_by_group.items():
            for i, _need in pairs:
                if group_left.get(g, 0) <= 0:
                    continue

                status = status_list[i]
                if status == "IGNORE":
                    continue

                row_flm = max(1.0, float(flm.iloc[i]))
                current_pred = float(pred[i])
                current_supply = max(0.0, float(qoh.iloc[i] + supply.iloc[i]) + current_pred)

                # Demand-protective ceiling.
                d60_i = float(d60.iloc[i])
                ttm_i = float(ttm.iloc[i])
                d30_i = float(d30.iloc[i])
                l30_i = float(l30.iloc[i])
                demand_ceiling_parts = []
                if d60_i > 0:
                    demand_ceiling_parts.append(d60_i + cfg.overstock_buffer_flm * row_flm)
                if ttm_i > 0:
                    demand_ceiling_parts.append(ttm_i / 3.0 + cfg.overstock_buffer_flm * row_flm)
                if d30_i > 0:
                    demand_ceiling_parts.append(d30_i * 1.25 + cfg.overstock_buffer_flm * row_flm)
                if l30_i > 0:
                    demand_ceiling_parts.append(l30_i * 1.75 + cfg.overstock_buffer_flm * row_flm)

                if demand_ceiling_parts:
                    ceiling = max(demand_ceiling_parts)
                else:
                    ceiling = current_supply + row_flm

                target_need = raw_need[i] * pass_multiplier
                desired_final_supply = max(current_supply, float(qoh.iloc[i] + supply.iloc[i]) + target_need)
                desired_final_supply = min(desired_final_supply, ceiling)
                add_need = max(0.0, desired_final_supply - current_supply)

                # Avoid weak Z-No-Alloc additions unless they have real demand.
                if status == "Z_NO_ALLOC" and raw_need[i] < row_flm * 0.40:
                    add_need = 0.0

                allocation = mround_to_flm(add_need, row_flm, group_left[g])
                if allocation <= 0:
                    continue

                pred[i] += allocation
                pass_allocs[pass_num][i] += allocation
                group_left[g] -= allocation

                if reasons[i]:
                    reasons[i] += "; "
                reasons[i] += f"Pass {pass_num}: +{allocation}"

    # Blank preservation:
    # - rows outside Allocate/Review/Z-No-Alloc remain blank
    # - zero predictions remain blank by default
    pred_series = pd.Series(pred, index=out.index)
    predicted_display = pred_series.astype(object)
    predicted_display[pred_series <= 0] = ""

    audit = pd.DataFrame({
        "Row": np.arange(2, n + 2),
        "Status": status_list,
        "Raw Need Score": np.round(raw_need, 3),
        "Predicted Final Alloc": predicted_display,
        "Pass 1 Add": pass_allocs[1],
        "Pass 2 Add": pass_allocs[2],
        "Pass 3 Add": pass_allocs[3],
        "Reason": reasons,
    })

    # Add useful source fields to audit when present.
    for label, key in [
        ("Flag", "flag"),
        ("Item/Group", "group_key"),
        ("FLM", "flm"),
        ("Left DC", "left_dc"),
        ("Proj. Demand", "proj_demand"),
        ("Alloc. Rec.", "alloc_rec"),
        ("D60", "d60"),
        ("TTM", "ttm"),
        ("QOH", "qoh"),
        ("Supply", "supply"),
    ]:
        col = cmap.get(key)
        if col and col in out.columns:
            audit[label] = out[col].values

    # Update Final Alloc column or create one if missing.
    if final_col and final_col in out.columns:
        out[final_col] = predicted_display
    else:
        out["Final Alloc."] = predicted_display

    # Update/derive Final Supply where present: AP + AQ + predicted allocation.
    if final_supply_col and final_supply_col in out.columns:
        out[final_supply_col] = (qoh + supply + pred_series).round(0).astype(int)

    return out, audit


def dataframe_to_xlsx_bytes(result_df: pd.DataFrame, audit_df: pd.DataFrame, summary: Dict) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="Allocation Output")
        audit_df.to_excel(writer, index=False, sheet_name="Allocation Audit")
        pd.DataFrame([summary]).to_excel(writer, index=False, sheet_name="Run Summary")

        wb = writer.book
        for ws in wb.worksheets:
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                values = [str(c.value) if c.value is not None else "" for c in col_cells[:200]]
                width = min(max(10, int(np.percentile([len(v) for v in values], 90)) + 2), 45)
                ws.column_dimensions[col_letter].width = width
    output.seek(0)
    return output.getvalue()


def dataframe_to_csv_bytes(result_df: pd.DataFrame) -> bytes:
    return result_df.to_csv(index=False).encode("utf-8")


def profile_dataframe(df: pd.DataFrame, cmap: Dict[str, Optional[str]]) -> Dict:
    final_col = cmap.get("final_alloc")
    flag_col = cmap.get("flag")
    left_col = cmap.get("left_dc")
    out = {
        "rows": int(len(df)),
        "columns": int(len(df.columns)),
        "detected_final_alloc_col": final_col or "created Final Alloc.",
        "detected_flag_col": flag_col or "not detected",
        "detected_left_dc_col": left_col or "not detected",
    }
    if flag_col:
        flags = safe_str_series(df[flag_col]).str.upper()
        out["allocate_rows"] = int(flags.str.contains("ALLOC", na=False).sum())
        out["review_rows"] = int(flags.str.contains("REVIEW", na=False).sum())
        out["z_no_alloc_rows"] = int((flags.str.contains("NO", na=False) & flags.str.contains("ALLOC", na=False)).sum())
    return out


# ============================================================
# UI
# ============================================================

st.title("📦 Allocation AI — Base Model v1/v2")
st.caption(APP_VERSION)

with st.sidebar:
    st.header("Model Selection")
    model_name = st.selectbox("Choose model", list(MODEL_CONFIGS.keys()), index=1)
    cfg = MODEL_CONFIGS[model_name]
    st.info(cfg.description)

    st.header("Output Options")
    output_format = st.radio("Download format", ["XLSX with audit", "CSV only"], index=0)
    show_preview_rows = st.slider("Preview rows", 10, 200, 50, step=10)

    st.header("Model Notes")
    st.markdown(
        """
        **Base Model v1**
        - More conservative
        - Better when over-allocation risk matters most
        - Closer to the Version 6 baseline

        **Base Model v2**
        - More complete demand blend
        - Stronger use of BM/BN
        - More willing to allocate Z - No Alloc opportunity rows
        - Closer to the Version 8 improvement
        """
    )

st.markdown(
    """
Upload an allocation file and the app will fill **Final Alloc.** using the selected base model.

**Supported inputs:** `.csv`, `.xlsx`, `.xlsm`, `.xlsb`  
**Main behavior:** fills Final Alloc only, keeps non-candidate rows blank, uses three allocation passes, rounds to FLM, and allows remaining DC units below one FLM.
"""
)

uploaded = st.file_uploader("Upload allocation file", type=["csv", "xlsx", "xlsm", "xlsb"])

if uploaded is None:
    st.warning("Upload a file to run the allocation model.")
    st.stop()

sheet_names = []
try:
    sheet_names = get_sheet_names(uploaded)
except Exception:
    sheet_names = []

selected_sheet = None
if sheet_names:
    selected_sheet = st.selectbox("Sheet", sheet_names, index=0)

try:
    df, meta = read_uploaded_file(uploaded, selected_sheet)
except Exception as e:
    st.error(f"Could not read the uploaded file: {e}")
    st.stop()

# Drop fully blank rows but preserve column order.
df = df.dropna(how="all").reset_index(drop=True)

cmap = build_column_map(df)
profile = profile_dataframe(df, cmap)

st.subheader("Detected File Structure")
col1, col2, col3, col4 = st.columns(4)
col1.metric("Rows", f"{profile['rows']:,}")
col2.metric("Columns", f"{profile['columns']:,}")
col3.metric("Flag Column", profile["detected_flag_col"])
col4.metric("Final Alloc Column", profile["detected_final_alloc_col"])

with st.expander("Detected column map", expanded=False):
    st.json({k: (v if v is not None else "not detected") for k, v in cmap.items()})

missing_critical = []
for key in ["flag", "left_dc", "flm"]:
    if not cmap.get(key):
        missing_critical.append(key)

if missing_critical:
    st.warning(
        "Some expected columns were not detected: "
        + ", ".join(missing_critical)
        + ". The app will use safe fallbacks where possible, but results may be less accurate."
    )

run = st.button("Run Allocation Model", type="primary")

if not run:
    st.stop()

with st.spinner("Running allocation model..."):
    result_df, audit_df = compute_model_predictions(df, cmap, cfg)

allocated_count = int((pd.to_numeric(audit_df["Predicted Final Alloc"], errors="coerce").fillna(0) > 0).sum())
total_units = int(pd.to_numeric(audit_df["Predicted Final Alloc"], errors="coerce").fillna(0).sum())
pass1 = int(audit_df["Pass 1 Add"].sum())
pass2 = int(audit_df["Pass 2 Add"].sum())
pass3 = int(audit_df["Pass 3 Add"].sum())

summary = {
    "app_version": APP_VERSION,
    "model": model_name,
    "run_timestamp": datetime.now().isoformat(timespec="seconds"),
    "input_file": uploaded.name,
    "sheet": selected_sheet or "",
    "rows": len(df),
    "allocated_rows": allocated_count,
    "total_units_allocated": total_units,
    "pass_1_units": pass1,
    "pass_2_units": pass2,
    "pass_3_units": pass3,
}

st.success("Allocation complete.")

m1, m2, m3, m4 = st.columns(4)
m1.metric("Allocated Rows", f"{allocated_count:,}")
m2.metric("Total Units", f"{total_units:,}")
m3.metric("Pass 1 / 2 / 3", f"{pass1:,} / {pass2:,} / {pass3:,}")
m4.metric("Model", model_name)

st.subheader("Allocation Audit Preview")
preview_cols = [
    c for c in [
        "Row", "Status", "Flag", "Item/Group", "FLM", "Left DC", "Proj. Demand",
        "Alloc. Rec.", "D60", "TTM", "QOH", "Supply", "Predicted Final Alloc",
        "Pass 1 Add", "Pass 2 Add", "Pass 3 Add", "Reason"
    ] if c in audit_df.columns
]
st.dataframe(audit_df[preview_cols].head(show_preview_rows), use_container_width=True)

st.subheader("Output Preview")
st.dataframe(result_df.head(show_preview_rows), use_container_width=True)

safe_model_name = model_name.lower().replace(" ", "_")
stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

if output_format == "XLSX with audit":
    out_bytes = dataframe_to_xlsx_bytes(result_df, audit_df, summary)
    st.download_button(
        "Download completed allocation workbook",
        data=out_bytes,
        file_name=f"allocation_output_{safe_model_name}_{stamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    out_bytes = dataframe_to_csv_bytes(result_df)
    st.download_button(
        "Download completed allocation CSV",
        data=out_bytes,
        file_name=f"allocation_output_{safe_model_name}_{stamp}.csv",
        mime="text/csv",
    )

with st.expander("Run summary", expanded=False):
    st.json(summary)
